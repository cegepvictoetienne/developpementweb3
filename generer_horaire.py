import openpyxl

fichier = "horaire.xlsx"

excel = openpyxl.load_workbook(fichier, data_only=True)

feuille = excel["horaire"]

with open("./wiki/horaire.md", "w") as f:
    # Écrire le titre de la page
    f.write("# Horaire du cours de développement web 3\n")

    # Écrire les entêtes du tableau de l'horaire
    entete = ""
    trait = ""
    for col in range(1, feuille.max_column + 1):
        entete += feuille.cell(row=1, column=col).value
        trait += "--"
        if col < feuille.max_column:
            entete += "|"
            trait += "|"
        else:
            entete += "\n"
            trait += "\n"

    f.write(entete)
    f.write(trait)

    for row in range(2, feuille.max_row + 1):
        ligne = ""
        for col in range(1, feuille.max_column + 1):
            valeur = feuille.cell(row=row, column=col).value
            if valeur is None:
                valeur = ""
            else:
                valeur = str(valeur)
            ligne += valeur
            if col < feuille.max_column:
                ligne += "|"
            else:
                ligne += "\n"

        f.write(ligne)

